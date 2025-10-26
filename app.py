"""
DeerLens – Streamlit front-end (disk-backed thumbnails + batching + pagination)
- Upload up to 900 JPEGs
- Runs Roboflow workflow asynchronously
- Shows KPI cards, stacked bar, time-of-day heat-map
- Clickable thumbnails with full-size viewer (modal if available)
- Inline editor (optional toggle) + Data Editor table (primary)
- Thumbnails stored on disk (temp dir) to keep RAM stable on Streamlit Cloud
"""

from __future__ import annotations

import asyncio
import base64  # still used for CSV building, not for images
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import os
import io
import gc
import uuid
import shutil
import tempfile
import aiohttp
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.colors import sequential
import streamlit as st
from io import BytesIO
from PIL import Image

from exif_utils import extract_datetime_original
from roboflow_client import RoboflowClient

# ──────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────
DIRECTIONS = ["", "N", "NE", "E", "SE", "S", "SW", "W", "NW"]  # "" = no entry
THUMB_MAX = (512, 512)  # hard cap to keep memory down
THUMB_QUAL = 75         # JPEG quality
DEFAULT_BATCH = 25      # process 20–50 at a time
DEFAULT_CONCURRENCY = 3 # aiohttp concurrency
THUMB_DISPLAY_W = 800   # UI width; actual file is ≤512px so browser scales up

# ──────────────────────────────────────────────────────────────────────
# Data structures
# ──────────────────────────────────────────────────────────────────────
@dataclass
class ImageResult:
    file_name: str
    date_time: Optional[str]
    buck_count: int
    deer_count: int
    doe_count: int
    annotated_path: str         # disk path to annotated thumbnail
    direction: Optional[str] = None
    target_buck: bool = False
    error: Optional[str] = None


# ──────────────────────────────────────────────────────────────────────
# Async inference helpers
# ──────────────────────────────────────────────────────────────────────
async def _process_single(
    session: aiohttp.ClientSession,
    rf_client: RoboflowClient,
    file_name: str,
    image_bytes: bytes,
    temp_dir: str,
) -> ImageResult:
    """Send one image to Roboflow and return counts + path to annotated JPEG on disk."""
    try:
        date_time = extract_datetime_original(image_bytes)
    except Exception:
        date_time = None

    try:
        annotated, counts = await rf_client.process_image(session, image_bytes, file_name)

        # shrink annotated JPEG aggressively to save RAM/IO
        try:
            img = Image.open(BytesIO(annotated))
            img.thumbnail(THUMB_MAX, Image.LANCZOS)
            buf = BytesIO()
            img.save(buf, format="JPEG", quality=THUMB_QUAL, optimize=True)
            annotated = buf.getvalue()
            try:
                img.close()
            except Exception:
                pass
        except Exception:
            # Fall back to original returned bytes
            pass

        # write to disk; keep only path in memory
        safe_name = f"{uuid.uuid4().hex}_{os.path.basename(file_name)}".replace(" ", "_")
        out_path = os.path.join(temp_dir, safe_name)
        with open(out_path, "wb") as f:
            f.write(annotated)

        return ImageResult(
            file_name=file_name,
            date_time=date_time,
            buck_count=counts.get("buck", 0),
            deer_count=counts.get("deer", 0),
            doe_count=counts.get("doe", 0),
            annotated_path=out_path,
            target_buck=False,
        )
    except Exception as exc:
        return ImageResult(
            file_name=file_name,
            date_time=date_time,
            buck_count=0,
            deer_count=0,
            doe_count=0,
            annotated_path="",
            target_buck=False,
            error=str(exc),
        )


async def process_images_async(
    files: List[Tuple[str, bytes]], rf_client: RoboflowClient, temp_dir: str
) -> List[ImageResult]:
    """Concurrently process many images (but only those provided in this batch)."""
    results, processed, total = [], 0, len(files)
    bar = st.progress(0)
    sem = asyncio.Semaphore(DEFAULT_CONCURRENCY)

    connector = aiohttp.TCPConnector(limit=DEFAULT_CONCURRENCY)
    async with aiohttp.ClientSession(connector=connector) as session:

        async def task(name: str, data: bytes) -> ImageResult:
            async with sem:
                try:
                    return await _process_single(session, rf_client, name, data, temp_dir)
                finally:
                    # free raw upload bytes ASAP
                    del data

        tasks = [task(n, b) for n, b in files]
        for fut in asyncio.as_completed(tasks):
            res = await fut
            results.append(res)
            processed += 1
            bar.progress(processed / max(total, 1))
    bar.empty()
    gc.collect()
    return results


# ──────────────────────────────────────────────────────────────────────
# Data wrangling
# ──────────────────────────────────────────────────────────────────────
def to_dataframe(results: List[ImageResult]) -> pd.DataFrame:
    """Results ➜ tidy DataFrame (small metadata only)."""
    return pd.DataFrame(
        [
            dict(
                file_name=r.file_name,
                date_time=r.date_time,
                buck_count=r.buck_count,
                deer_count=r.deer_count,
                doe_count=r.doe_count,
                direction=r.direction,
                target_buck=bool(r.target_buck),
            )
            for r in results
        ]
    )


def compute_summary(df: pd.DataFrame) -> Dict[str, int]:
    return dict(
        total_images=len(df),
        total_buck=int(df["buck_count"].sum()),
        total_deer=int(df["deer_count"].sum()),
        total_doe=int(df["doe_count"].sum()),
    )


def bucket_time(ts: pd.Timestamp | None) -> str | None:
    if pd.isna(ts):
        return None
    m = ts.hour * 60 + ts.minute
    if 360 <= m <= 480:
        return "Dawn"
    if 481 <= m <= 660:
        return "Morning"
    if 661 <= m <= 960:
        return "Midday"
    if 961 <= m <= 1110:
        return "Afternoon"
    if 1111 <= m <= 1259:
        return "Evening"
    return "Night"


def categorise(results: List[ImageResult], df_counts: pd.DataFrame) -> Dict[str, List[ImageResult]]:
    """Bucket each ImageResult according to edited counts."""
    latest = df_counts.set_index("file_name")
    cat: Dict[str, List[ImageResult]] = {"Buck": [], "Deer": [], "Doe": [], "No Tag": []}
    for res in results:
        if res.file_name not in latest.index:
            cat["No Tag"].append(res)
            continue
        row = latest.loc[res.file_name]
        counts = {"Buck": row.buck_count, "Deer": row.deer_count, "Doe": row.doe_count}
        mx = max(counts.values())
        if mx == 0:
            cat["No Tag"].append(res)
        elif counts["Buck"] == mx:
            cat["Buck"].append(res)
        elif counts["Deer"] == mx:
            cat["Deer"].append(res)
        else:
            cat["Doe"].append(res)
    return cat


def cleanup_temp_dir():
    td = st.session_state.get("temp_dir")
    if td and os.path.isdir(td):
        try:
            shutil.rmtree(td, ignore_errors=True)
        except Exception:
            pass
    st.session_state.pop("temp_dir", None)


# ──────────────────────────────────────────────────────────────────────
# Streamlit UI
# ──────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="DeerLens", layout="wide")
st.title("🦌 DeerLens – Deer Detection and Analysis")
st.markdown(
    """
Upload between **1 and 900** JPEG images from your trail cam.
DeerLens detects deer, draws bounding boxes, and gives you instant stats.
"""
)

uploader = st.file_uploader("Select JPEG images", type=["jpg", "jpeg"], accept_multiple_files=True)

# Optional viewer fallback (if st.modal not available)
if st.session_state.get("viewer_path") and not hasattr(st, "modal"):
    with st.expander(f"Full-size viewer: {os.path.basename(st.session_state.get('viewer_name','image'))}", expanded=True):
        st.image(st.session_state["viewer_path"], use_container_width=True)
        if st.button("Close viewer"):
            st.session_state.pop("viewer_path", None)
            st.session_state.pop("viewer_name", None)

# Run pipeline ---------------------------------------------------------
if uploader:
    if not 1 <= len(uploader) <= 900:
        st.warning("Select 1–900 images.")
        st.stop()

    col_btn, col_clear = st.columns([1, 1])
    with col_btn:
        go_run = st.button("Process Images", type="primary")
    with col_clear:
        if st.button("Clear cached images"):
            cleanup_temp_dir()
            st.success("Cleared temporary thumbnails.")

    if go_run:
        # Clean previous run's temp files
        cleanup_temp_dir()
        st.session_state["temp_dir"] = tempfile.mkdtemp(prefix="deerlens_")

        try:
            client = RoboflowClient()
        except Exception as e:
            st.error(str(e))
            st.stop()

        total = len(uploader)
        BATCH = DEFAULT_BATCH
        bar = st.progress(0.0, text="Processing images…")
        all_results: List[ImageResult] = []

        try:
            for i in range(0, total, BATCH):
                batch = uploader[i:i + BATCH]
                files_data = [(f.name, f.read()) for f in batch]
                batch_results = asyncio.run(process_images_async(files_data, client, st.session_state["temp_dir"]))
                all_results.extend(batch_results)
                files_data.clear()
                gc.collect()
                bar.progress((i + len(batch)) / total, text=f"{i + len(batch)}/{total} processed")
            bar.empty()
        except MemoryError:
            st.error("Ran out of memory while processing. Try a smaller batch size (e.g., 20).")
            st.stop()
        except Exception as e:
            st.exception(e)
            st.stop()

        # Cache results for UI / editing
        st.session_state["image_results"] = all_results
        st.session_state["orig_df"] = to_dataframe(all_results)
        st.session_state["edited_df"] = st.session_state["orig_df"].copy()
        if "open_cat" not in st.session_state:
            st.session_state.open_cat = None

# Guard: only proceed if results exist
if "image_results" not in st.session_state or "edited_df" not in st.session_state:
    st.info("Upload images and click **Process Images** to see results.")
    st.stop()

results = st.session_state["image_results"]

# Editable grid (primary editing surface) ------------------------------
st.subheader("Review / correct counts (table editor)")
edited = st.data_editor(
    st.session_state["edited_df"],
    disabled=["file_name", "date_time"],
    key="editor",
    num_rows="fixed",
    use_container_width=True,
)
if st.button("Apply overrides from table", key="apply_table"):
    st.session_state["edited_df"] = edited

# KPI/Charts/Downloads containers
kpi_container = st.container()
charts_container = st.container()
download_container = st.container()

# Remember last category pane
if "open_cat" not in st.session_state:
    st.session_state.open_cat = None

# Gallery controls -----------------------------------------------------
st.header("Annotated Images")

# Images per page = total across categories; we’ll show ~¼ per category
PAGE_SIZE = st.selectbox("Images per page (total, across all categories)", [12, 24, 48], index=1)
PER_CAT = max(1, PAGE_SIZE // 4)
page = st.number_input("Page", min_value=1, value=1, step=1)
show_inline = st.checkbox("Enable per-image inline overrides (slower)", value=False)

cats = categorise(results, st.session_state["edited_df"])

def _sort_key_datetime(value):
    dt = pd.to_datetime(value, errors="coerce")
    return (pd.isna(dt), dt if pd.notna(dt) else pd.Timestamp.max)

# ▒▒▒ bulk-edit form ▒▒▒  (only commits inline widgets if enabled)
with st.form("bulk_overrides", clear_on_submit=False):
    for cat_name in ["Buck", "Deer", "Doe", "No Tag"]:
        imgs = cats.get(cat_name, [])

        # sort by date_time (earliest first) and paginate per category
        imgs = sorted(imgs, key=lambda r: _sort_key_datetime(
            st.session_state["edited_df"].loc[
                st.session_state["edited_df"]["file_name"] == r.file_name, "date_time"
            ].iloc[0] if (st.session_state["edited_df"]["file_name"] == r.file_name).any() else None
        ))
        start = (page - 1) * PER_CAT
        end = start + PER_CAT
        total_in_cat = len(imgs)
        imgs = imgs[start:end]

        expanded_now = st.session_state.open_cat == cat_name
        with st.expander(f"{cat_name} ({total_in_cat}) • showing {len(imgs)} on this page", expanded=expanded_now):
            if st.session_state.open_cat != cat_name:
                st.session_state.open_cat = cat_name

            if not imgs:
                st.write("No images on this page.")
                continue

            for res in imgs:
                row = st.session_state["edited_df"].loc[
                    st.session_state["edited_df"]["file_name"] == res.file_name
                ]
                row = row.iloc[0] if not row.empty else None

                col_img, col_edit = st.columns([3, 2], gap="small")

                with col_img:
                    path = res.annotated_path
                    if not path or not os.path.exists(path):
                        st.write(f"{res.file_name} (no annotated image)")
                    else:
                        st.image(path, width=THUMB_DISPLAY_W, caption=res.file_name, output_format="JPEG")
                        if st.button("🔍 Open full size", key=f"open_{res.file_name}"):
                            st.session_state["viewer_path"] = path
                            st.session_state["viewer_name"] = res.file_name
                        # Use modal if available
                        if hasattr(st, "modal") and st.session_state.get("viewer_path") == path:
                            with st.modal(res.file_name, key=f"modal_{res.file_name}"):
                                st.image(path, use_container_width=True)
                                if st.button("Close", key=f"close_{res.file_name}"):
                                    st.session_state.pop("viewer_path", None)
                                    st.session_state.pop("viewer_name", None)

                if show_inline and row is not None:
                    with col_edit:
                        st.markdown("**Manual overrides (inline)**")
                        dir_opts = DIRECTIONS[1:]
                        current_dir = row.direction if pd.notna(row.direction) else ""
                        st.number_input("Buck", min_value=0, value=int(row.buck_count), key=f"buck_{res.file_name}")
                        st.number_input("Deer", min_value=0, value=int(row.deer_count), key=f"deer_{res.file_name}")
                        st.number_input("Doe",  min_value=0, value=int(row.doe_count),  key=f"doe_{res.file_name}")
                        st.selectbox(
                            "Direction (optional)",
                            options=["—"] + dir_opts,
                            index=(dir_opts.index(current_dir) + 1) if current_dir else 0,
                            key=f"dir_{res.file_name}",
                        )
                        st.checkbox(
                            "Target Buck (optional)",
                            value=bool(row.target_buck) if pd.notna(row.target_buck) else False,
                            key=f"tb_{res.file_name}",
                        )

    save_all = st.form_submit_button(
        "Apply ALL inline overrides on this page",
        type="primary",
        disabled=not show_inline
    )

# Commit inline overrides (only for keys that exist) -------------------
if save_all and show_inline:
    ed_df = st.session_state["edited_df"]
    # Only update rows that had widgets on this page (keys present)
    for fname in ed_df["file_name"]:
        # For each field, only commit if the widget key exists to avoid zeroing others
        k_b, k_d, k_do, k_dir, k_tb = (f"buck_{fname}", f"deer_{fname}", f"doe_{fname}", f"dir_{fname}", f"tb_{fname}")
        updates = {}
        if k_b in st.session_state:   updates["buck_count"] = int(st.session_state[k_b])
        if k_d in st.session_state:   updates["deer_count"] = int(st.session_state[k_d])
        if k_do in st.session_state:  updates["doe_count"]  = int(st.session_state[k_do])
        if k_dir in st.session_state:
            dr = st.session_state[k_dir]
            updates["direction"] = None if dr == "—" else dr
        if k_tb in st.session_state:  updates["target_buck"] = bool(st.session_state[k_tb])
        if updates:
            ed_df.loc[ed_df["file_name"] == fname, list(updates.keys())] = list(updates.values())
    st.session_state["edited_df"] = ed_df
    st.rerun()

# ──────────────────────────────────────────────────────────────────
# 🔄  Refresh KPI, charts, and CSV  (runs after any edits) 🔄
# ──────────────────────────────────────────────────────────────────
df_final = st.session_state["edited_df"]
smry_final = compute_summary(df_final)

# ----- KPI cards ------------------------------------------------------
with kpi_container:
    st.header("Summary Metrics")
    k = st.columns(4)
    k[0].metric("Images", smry_final["total_images"])
    k[1].metric("Buck",   smry_final["total_buck"])
    k[2].metric("Deer",   smry_final["total_deer"])
    k[3].metric("Doe",    smry_final["total_doe"])

# ----- Charts ---------------------------------------------------------
with charts_container:
    # stacked bar
    if df_final["date_time"].notna().any():
        dt = pd.to_datetime(df_final["date_time"], errors="coerce")
        bar_df = df_final.copy()
        bar_df["date"] = dt.dt.date
        agg = (
            bar_df.groupby("date", as_index=False)
            .sum(numeric_only=True)
            .sort_values("date")
        )
        melt = agg.melt("date", ["buck_count", "deer_count", "doe_count"],
                        var_name="Class", value_name="Count")
        melt["Class"] = melt["Class"].str.replace("_count", "").str.title()
        fig_bar = px.bar(
            melt,
            x="date",
            y="Count",
            color="Class",
            color_discrete_map={"Buck": "#228B22", "Doe": "#FFC0CB", "Deer": "#D2B48C"},
            title="Activity by Date (stacked)",
            labels={"date": "Date"},
        )
        fig_bar.update_layout(barmode="stack", xaxis_tickangle=-45)
        st.plotly_chart(fig_bar, use_container_width=True)

    # heat-map + overlays (green = bucks, red = target-buck sightings)
    if df_final["date_time"].notna().any():
        ts = pd.to_datetime(df_final["date_time"], errors="coerce")
        heat_df = df_final.copy()
        heat_df["date"]   = ts.dt.date
        heat_df["bucket"] = ts.apply(bucket_time)

        if "target_buck" not in heat_df.columns:
            heat_df["target_buck"] = False
        heat_df["target_buck"] = heat_df["target_buck"].fillna(False).astype(bool)
        heat_df["tb_sightings"] = heat_df["target_buck"].astype(int)

        agg = (
            heat_df.groupby(["bucket", "date"], as_index=False)
            .agg(
                buck=("buck_count", "sum"),
                deer=("deer_count", "sum"),
                doe=("doe_count", "sum"),
                tb_sightings=("tb_sightings", "sum"),
            )
        )
        agg["activity"] = agg["buck"] + agg["deer"] + agg["doe"]

        buckets = ["Dawn", "Morning", "Midday", "Afternoon", "Evening", "Night"]
        dates   = sorted([d for d in agg["date"].unique() if pd.notna(d)])

        bucket_labels = {
            "Dawn":      "Dawn (6:00–8:00 am)",
            "Morning":   "Morning (8:01–11:00 am)",
            "Midday":    "Midday (11:01 am–4:00 pm)",
            "Afternoon": "Afternoon (4:01–6:30 pm)",
            "Evening":   "Evening (6:31–8:59 pm)",
            "Night":     "Night (9:00 pm–5:59 am)",
        }

        # Heat matrix
        z = [[int(agg.query("bucket==@b and date==@d")["activity"].sum() or 0)
              for d in dates] for b in buckets]

        # GREEN bubbles = total buck count
        sx, sy, ss, sc = [], [], [], []
        for b in buckets:
            for d in dates:
                bk = int(agg.query("bucket==@b and date==@d")["buck"].sum() or 0)
                if bk:
                    sx.append(d); sy.append(b); ss.append(bk * 10 + 8); sc.append(bk)

        # RED bubbles = number of target-buck sightings
        rsx, rsy, rss, rsc = [], [], [], []
        for b in buckets:
            for d in dates:
                tb = int(agg.query("bucket==@b and date==@d")["tb_sightings"].sum() or 0)
                if tb:
                    rsx.append(d); rsy.append(b); rss.append(tb * 8 + 6); rsc.append(tb)

        heat = go.Heatmap(
            z=z, x=dates, y=buckets,
            colorscale=sequential.Blues,
            colorbar=dict(title="Total Activity"),
            hovertemplate="Date %{x}<br>%{y}<br>Activity %{z}<extra></extra>",
        )
        dots_green = go.Scatter(
            name="Buck (count)",
            x=sx, y=sy, mode="markers",
            marker=dict(size=ss, color="#228B22", opacity=.85,
                        line=dict(width=1, color="white")),
            customdata=sc,
            hovertemplate="Date %{x}<br>%{y}<br>Bucks %{customdata}<extra></extra>",
            showlegend=True,
        )
        dots_red = go.Scatter(
            name="Target Buck (sightings)",
            x=rsx, y=rsy, mode="markers",
            marker=dict(size=rss, color="#C62828", opacity=.9,
                        line=dict(width=1, color="white")),
            customdata=rsc,
            hovertemplate="Date %{x}<br>%{y}<br>Target-Buck Sightings %{customdata}<extra></extra>",
            showlegend=True,
        )

        fig_hm = go.Figure([heat, dots_green, dots_red])
        fig_hm.update_yaxes(autorange="reversed")
        fig_hm.update_yaxes(tickmode="array", tickvals=buckets,
                            ticktext=[bucket_labels[b] for b in buckets])
        fig_hm.update_layout(
            title=("Heat-map of Total Activity<br>"
                   "<span style='font-size:0.8em'>green = buck count; red = # target-buck sightings</span>"),
            xaxis_tickangle=-45, xaxis_type="category",
            yaxis_title="Time of Day", xaxis_title="Date",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        )
        st.plotly_chart(fig_hm, use_container_width=True)

# ----- CSV & Excel / PNG downloads (Matplotlib fallback) ------------
with download_container:
    summary_df = pd.DataFrame(
        [
            {"metric": "total_images", "value": smry_final["total_images"]},
            {"metric": "total_buck",   "value": smry_final["total_buck"]},
            {"metric": "total_deer",   "value": smry_final["total_deer"]},
            {"metric": "total_doe",    "value": smry_final["total_doe"]},
        ]
    )

    # CSV (no images in CSV)
    buf = io.StringIO()
    summary_df.to_csv(buf, index=False)
    buf.write("\n")
    df_final.to_csv(buf, index=False)
    csv_bytes = buf.getvalue().encode("utf-8")

    st.download_button(
        "Download revised CSV (+ summary)",
        csv_bytes,
        "deerlens_results.csv",
        "text/csv",
    )

    # Build a PNG of the heatmap using Matplotlib (no Chrome/Kaleido required)
    png_bytes = None
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np

        DPI = 150
        GREEN_K, GREEN_B = 10.0, 8.0
        RED_K,   RED_B   =  8.0, 6.0
        def px_to_mpl_area(px: float, dpi: float = DPI) -> float:
            pts = px * (72.0 / dpi)
            return pts * pts

        ts = pd.to_datetime(df_final["date_time"], errors="coerce")
        heat_df = df_final.copy()
        heat_df["date"]   = ts.dt.date
        heat_df["bucket"] = ts.apply(bucket_time)

        if "target_buck" not in heat_df.columns:
            heat_df["target_buck"] = False
        heat_df["target_buck"]  = heat_df["target_buck"].fillna(False).astype(bool)
        heat_df["tb_sightings"] = heat_df["target_buck"].astype(int)

        agg = (
            heat_df.groupby(["bucket", "date"], as_index=False)
            .agg(
                buck=("buck_count", "sum"),
                deer=("deer_count", "sum"),
                doe=("doe_count", "sum"),
                tb_sightings=("tb_sightings", "sum"),
            )
        )
        agg["activity"] = agg["buck"] + agg["deer"] + agg["doe"]

        buckets = ["Dawn", "Morning", "Midday", "Afternoon", "Evening", "Night"]
        dates   = sorted([d for d in agg["date"].unique() if pd.notna(d)])

        z = np.array(
            [
                [int(agg.query("bucket==@b and date==@d")["activity"].sum() or 0)
                 for d in dates]
                for b in buckets
            ],
            dtype=float,
        )

        gx, gy, gs = [], [], []
        rx, ry, rs = [], [], []
        for bi, b in enumerate(buckets):
            for di, d in enumerate(dates):
                bk = int(agg.query("bucket==@b and date==@d")["buck"].sum() or 0)
                tb = int(agg.query("bucket==@b and date==@d")["tb_sightings"].sum() or 0)
                if bk:
                    size_px = GREEN_K * bk + GREEN_B
                    gx.append(di); gy.append(bi); gs.append(px_to_mpl_area(size_px, DPI))
                if tb:
                    size_px_r = RED_K * tb + RED_B
                    rx.append(di); ry.append(bi); rs.append(px_to_mpl_area(size_px_r, DPI))

        fig_mpl, ax = plt.subplots(figsize=(12, 6), dpi=DPI)
        im = ax.imshow(z, aspect="auto", cmap="Blues", origin="upper")
        if gx:
            ax.scatter(gx, gy, s=gs, marker="o", zorder=2,
                       facecolors="#228B22", edgecolors="white", linewidths=1, alpha=0.85)
        if rx:
            ax.scatter(rx, ry, s=rs, marker="o", zorder=3,
                       facecolors="#C62828", edgecolors="white", linewidths=1, alpha=0.9)
        ax.set_xticks(range(len(dates)))
        ax.set_xticklabels([str(d) for d in dates], rotation=45, ha="right")
        ax.set_yticks(range(len(buckets)))
        ax.set_yticklabels([
            "Dawn (6:00–8:00 am)",
            "Morning (8:01–11:00 am)",
            "Midday (11:01 am–4:00 pm)",
            "Afternoon (4:01–6:30 pm)",
            "Evening (6:31–8:59 pm)",
            "Night (9:00 pm–5:59 am)",
        ])
        ax.set_xlabel("Date")
        ax.set_ylabel("Time of Day")
        ax.set_title("Heat-map of Total Activity (green=buck count, red=# target-buck sightings)")
        cbar = plt.colorbar(im, ax=ax)
        cbar.set_label("Total Activity")
        ax.set_xlim(-0.5, max(len(dates) - 0.5, 0.5))
        ax.set_ylim(-0.5, len(buckets) - 0.5)

        plt.tight_layout()
        _buf = BytesIO()
        fig_mpl.savefig(_buf, format="png", bbox_inches="tight")
        plt.close(fig_mpl)
        png_bytes = _buf.getvalue()
    except Exception as e:
        png_bytes = None
        st.warning(f"Could not render heatmap PNG fallback: {e}")

    # Excel with optional Heatmap sheet
    xlsx_buf = BytesIO()
    try:
        import xlsxwriter  # if present, prefer it
        excel_engine = "xlsxwriter"
    except Exception:
        excel_engine = "openpyxl"

    with pd.ExcelWriter(xlsx_buf, engine=excel_engine) as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        df_final.to_excel(writer, index=False, sheet_name="Data")

        if png_bytes:
            if excel_engine == "xlsxwriter":
                ws = writer.book.add_worksheet("Heatmap")
                writer.sheets["Heatmap"] = ws
                ws.set_column("A:A", 2)
                ws.insert_image(
                    "B2",
                    "heatmap.png",
                    {"image_data": BytesIO(png_bytes), "x_scale": 1.0, "y_scale": 1.0},
                )
            else:
                from openpyxl.drawing.image import Image as XLImage
                from PIL import Image as PILImage
                ws = writer.book.create_sheet("Heatmap")
                writer.sheets["Heatmap"] = ws
                pil_img = PILImage.open(BytesIO(png_bytes))
                xl_img = XLImage(pil_img)
                ws.add_image(xl_img, "B2")
        else:
            if excel_engine == "xlsxwriter":
                ws = writer.book.add_worksheet("Heatmap")
                writer.sheets["Heatmap"] = ws
                ws.write("A1", "Heatmap image not available.")
            else:
                ws = writer.book.create_sheet("Heatmap")
                writer.sheets["Heatmap"] = ws
                ws["A1"] = "Heatmap image not available."

    xlsx_data = xlsx_buf.getvalue()
    st.download_button(
        "Download Excel (data + Heatmap image)",
        xlsx_data,
        "deerlens_results.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
